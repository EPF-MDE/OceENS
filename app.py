"""
=============================================================================
OceENS - Application principale FastAPI (version fusionnée)
=============================================================================
Combine :
- L'authentification Azure Entra ID (auth.py)
- La gestion des sessions (SessionMiddleware)
- Les routes du module app (1).py (sondages, questionnaires, API)
- Les dashboards par rôle
"""

from dotenv import load_dotenv


import os
import io
from typing import Annotated, Dict, List, Optional
from datetime import datetime
from contextlib import asynccontextmanager

from fastapi import Depends, FastAPI, File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlmodel import Session, SQLModel, create_engine, select, delete, func
import uvicorn

# ┌─ Importation des modèles et du module d'authentification ─────────────┐
from models import (
    Module,
    Question,
    Repondre,
    Reponse,
    Section,
    Sondage,
    Template,
    Option,
    User,
)
from pydantic import BaseModel
from starlette.middleware.sessions import SessionMiddleware
from auth import router as auth_router, get_current_user, require_roles
from sondage_loader import load_sondage_complet
from services.export_csv import generate_csv_response
from services.visualisation_data import get_visualisation_context

load_dotenv()
# ┌─ Configuration ────────────────────────────────────────────────────────┐
# Les trois slugs de dashboard reconnus par l'application
VALID_ROLES = {"admin", "etudiant", "rprm"}


def role_to_dashboard_slug(role: str) -> str:
    """
    Convertit le rôle stocké en BDD en slug de route dashboard.

    "Admin"              → "admin"
    "RP-RM"              → "rprm"
    "RP-RM:MDE_P2027"    → "rprm"
    "Etudiant" (ou autre) → "etudiant"
    """
    if role.startswith("Admin"):
        return "admin"
    elif role.startswith("RP-RM"):
        return "rprm"
    else:
        return "etudiant"


def parse_rprm_formations(role: str) -> list[str]:
    """
    Extrait la liste des formations autorisées depuis une chaîne de rôle RP-RM.

    "RP-RM:FORMATION1;FORMATION2" → ["FORMATION1", "FORMATION2"]
    "RP-RM:FORMATION1"            → ["FORMATION1"]
    "RP-RM"                       → []
    "Admin"                       → []
    "Admin:FORMATION1;FORMATION2" → ["FORMATION1", "FORMATION2"]
    "Admin:FORMATION1"            → ["FORMATION1"]
    """
    if not role or not isinstance(role, str):
        return []
    role_upper = role.strip()
    if not (role_upper.startswith("RP-RM:") or role_upper.startswith("Admin:")):
        return []
    after_colon = role_upper.split(":", 1)[1]
    return [f.strip() for f in after_colon.split(";") if f.strip()]


# └────────────────────────────────────────────────────────────────────────┘


# ┌─ Configuration de la base de données ──────────────────────────────────┐
sqlite_file_name = "database/db_oceens.db"
sqlite_url = f"sqlite:///{sqlite_file_name}"

connect_args = {"check_same_thread": False, "timeout": 15}
engine = create_engine(sqlite_url, connect_args=connect_args)


def create_db_and_tables():
    SQLModel.metadata.create_all(engine)


def get_session():
    with Session(engine) as session:
        yield session


SessionDep = Annotated[Session, Depends(get_session)]
# └────────────────────────────────────────────────────────────────────────┘


# ┌─ Modèles Pydantic pour les données entrantes ────────────────────────┐
class SondageCreate(BaseModel):
    id_template: int
    campus: str
    formation: str
    semestre: str
    annee_scolaire: str
    id_user: Optional[int] = 1


class ProfesseurBase(BaseModel):
    id: int
    prenom: str
    nom: str


class ModuleCreate(BaseModel):
    id: int
    nom: str
    choix_enseignant_exclusif: bool = False
    professeurs: List[ProfesseurBase]


class UECreate(BaseModel):
    id: int
    nom: str
    optionnel: bool
    modules: List[ModuleCreate]


class SondageFullCreate(BaseModel):
    id_template: int
    campus: str
    formation: str
    semestre: str
    annee_scolaire: str
    ues: List[UECreate]


class ReponseItem(BaseModel):
    id_section: int
    id_question: int
    valeur: str
    module_id: Optional[int] = None
    enseignant: Optional[str] = None


class QuestionnaireSubmission(BaseModel):
    reponses: List[ReponseItem]


class RoleUpdate(BaseModel):
    role: str


import json

# └────────────────────────────────────────────────────────────────────────┘


# ┌─ Fonctions utilitaires ──────────────────────────────────────────────┐
def parse_name(full_name: Optional[str], fallback_id: int) -> Dict[str, Optional[str]]:
    if not full_name:
        return {"id": fallback_id, "prenom": None, "nom": None}
    parts = full_name.strip().split()
    if len(parts) == 1:
        return {"id": fallback_id, "prenom": parts[0], "nom": ""}
    return {"id": fallback_id, "prenom": parts[0], "nom": " ".join(parts[1:])}


def build_parametrage_data(
    session: Session, allowed_formations: list[str] | None = None
) -> Dict[str, object]:
    templates = session.exec(select(Template)).all()
    sondages = session.exec(select(Sondage)).all()
    modules = session.exec(select(Module)).all()
    users = session.exec(select(User)).all()

    campus_names = []
    filiere_names = []
    semestres = []
    annees_scolaires = []
    formation_to_campus: Dict[str, str] = {}
    for sondage in sondages:
        if sondage.campus and sondage.campus not in campus_names:
            campus_names.append(sondage.campus)
        if sondage.formation and sondage.formation not in filiere_names:
            filiere_names.append(sondage.formation)
            formation_to_campus[sondage.formation] = sondage.campus or ""
        if sondage.semestre and sondage.semestre not in semestres:
            semestres.append(sondage.semestre)
        if sondage.annee_scolaire and sondage.annee_scolaire not in annees_scolaires:
            annees_scolaires.append(sondage.annee_scolaire)

    # Filtrer les filières si l'utilisateur RP-RM n'a accès qu'à certaines formations
    if allowed_formations is not None:
        filiere_names = [f for f in filiere_names if f in allowed_formations]
        # Ajouter les formations autorisées absentes des sondages existants
        for af in allowed_formations:
            if af not in filiere_names:
                filiere_names.append(af)

    default_campuses = ["Paris-Cachan", "Montpellier", "Troyes", "St-Nazaire"]
    for dc in default_campuses:
        if dc not in campus_names:
            campus_names.append(dc)

    campus_list = [
        {"id": index + 1, "nom": campus} for index, campus in enumerate(campus_names)
    ]
    campus_index = {campus["nom"]: campus["id"] for campus in campus_list}
    filieres = []
    for index, formation in enumerate(filiere_names):
        filieres.append(
            {
                "id": index + 1,
                "nom": formation,
                "campus_id": campus_index.get(
                    formation_to_campus.get(formation, ""), None
                ),
            }
        )

    professors = []
    professor_index = 1
    seen_professors = {}
    for module in modules:
        if not module.enseignant:
            continue
        prof_strings = [p.strip() for p in module.enseignant.split(",") if p.strip()]
        for prof_str in prof_strings:
            professor = parse_name(prof_str, professor_index)
            if not professor["prenom"] and not professor["nom"]:
                continue
            key = (professor["prenom"].lower(), professor["nom"].lower())
            if key not in seen_professors:
                seen_professors[key] = professor_index
                professor["id"] = professor_index
                professors.append(professor)
                professor_index += 1

    for user in users:
        if user.role and "Enseignant" in user.role and user.mail:
            parsed = parse_name(
                user.mail.split("@")[0].replace(".", " "), professor_index
            )
            parsed["nom"] = parsed["nom"] or ""
            parsed["prenom"] = parsed["prenom"] or ""
            key = (parsed["prenom"].lower(), parsed["nom"].lower())
            if key and key not in seen_professors:
                seen_professors[key] = professor_index
                parsed["id"] = professor_index
                professors.append(parsed)
                professor_index += 1

    ues_by_filiere = {}
    if modules and filieres:
        default_filiere_id = filieres[0]["id"]
        for module in modules:
            filiere_id = default_filiere_id
            ue_name = module.ue or "Sans UE"
            ues_by_filiere.setdefault(filiere_id, [])
            ue_entry = next(
                (ue for ue in ues_by_filiere[filiere_id] if ue["nom"] == ue_name), None
            )
            if ue_entry is None:
                ue_entry = {
                    "id": len(ues_by_filiere[filiere_id]) + 1,
                    "nom": ue_name,
                    "optionnel": module.ue_optionnelle or False,
                    "_open": True,
                    "modules": [],
                }
                ues_by_filiere[filiere_id].append(ue_entry)
            prof_list = []
            if module.enseignant:
                prof_strings = [
                    p.strip() for p in module.enseignant.split(",") if p.strip()
                ]
                for prof_str in prof_strings:
                    parsed = parse_name(prof_str, 0)
                    if parsed["prenom"] or parsed["nom"]:
                        key = (parsed["prenom"].lower(), parsed["nom"].lower())
                        if key not in seen_professors:
                            seen_professors[key] = professor_index
                            parsed["id"] = professor_index
                            professors.append(parsed)
                            professor_index += 1
                        prof_list.append(
                            {
                                "id": seen_professors[key],
                                "prenom": parsed["prenom"],
                                "nom": parsed["nom"],
                            }
                        )
            ue_entry["modules"].append(
                {
                    "id": int(module.id_module or 0),
                    "nom": module.nom or "Module",
                    "choix_enseignant_exclusif": bool(module.choix_enseignant),
                    "professeurs": prof_list,
                }
            )

    template_dicts = [template.dict() for template in templates]

    return {
        "templates": template_dicts,
        "campusList": campus_list,
        "filieres": filieres,
        "semestres": semestres,
        "anneesScolaires": annees_scolaires,
        "profsList": professors,
        "uesByFiliere": ues_by_filiere,
        "selectedTemplateId": template_dicts[0]["id_template"]
        if template_dicts
        else None,
        "selectedCampusId": campus_list[0]["id"] if campus_list else None,
        "selectedFiliereId": filieres[0]["id"] if filieres else None,
        "semestreAnnee": semestres[0] if semestres else "",
        "selectedAnneeScolaire": annees_scolaires[0] if annees_scolaires else "",
        "questions": [
            question.dict() for question in session.exec(select(Question)).all()
        ],
        "options": [option.dict() for option in session.exec(select(Option)).all()],
    }


# └────────────────────────────────────────────────────────────────────────┘


# ┌─ Gestion du cycle de vie (lifespan) ──────────────────────────────────┐
@asynccontextmanager
async def lifespan(app: FastAPI):
    print("Initialisation de la base de données...")
    create_db_and_tables()
    yield
    print("Fermeture de la connexion...")


# └────────────────────────────────────────────────────────────────────────┘


def create_app():
    """
    Crée et configure l'application FastAPI fusionnée.
    """
    app = FastAPI(
        title="OceENS",
        description="Système de gestion et de connexion pour étudiants, professeurs et admins",
        lifespan=lifespan,
    )

    # SessionMiddleware (authentification)
    app.add_middleware(
        SessionMiddleware,
        secret_key=os.environ.get("SECRET_KEY", "changeme"),
        https_only=True,
        same_site="lax",
    )

    # Routeur d'authentification (login/logout/callback Azure Entra ID)
    app.include_router(auth_router)

    # Fichiers statiques et templates (montés une seule fois)
    app.mount("/static", StaticFiles(directory="static"), name="static")
    templates = Jinja2Templates(directory="templates")

    # ┌─ Route : Page d'accueil (version app.py conservée) ──────────────┐
    @app.get("/", response_class=HTMLResponse)
    async def index(request: Request):
        """
        Page d'accueil. Si l'utilisateur est déjà connecté avec un rôle
        valide, redirection vers son dashboard. Sinon, affichage du login.
        """
        user = get_current_user(request)
        if user and user.get("role"):
            slug = role_to_dashboard_slug(user["role"])
            return RedirectResponse(url=f"/dashboard/{slug}")
        return templates.TemplateResponse(request=request, name="index.html")

    # └────────────────────────────────────────────────────────────────┘

    # ┌─ Route : Paramétrage (accès restreint Admin + RP-RM) ──────────────┐
    @app.get("/parametrage", response_class=HTMLResponse)
    def parametrage(request: Request, session: SessionDep):
        # ── Sécurité : vérifier que l'utilisateur est Admin ou RP-RM ──
        user = require_roles(request, ["Admin", "RP-RM"])
        if user is None:
            # Utilisateur non connecté ou rôle insuffisant → redirection
            connected_user = get_current_user(request)
            if connected_user:
                slug = role_to_dashboard_slug(connected_user.get("role", ""))
                return RedirectResponse(url=f"/dashboard/{slug}")
            return RedirectResponse(url="/")

        # Déterminer les formations autorisées pour un RP-RM
        allowed_formations = None
        is_rprm = False
        role = user.get("role", "") or ""

        if role.startswith("Admin:"):
            allowed_formations = parse_rprm_formations(role)
            is_rprm = True
        if role.startswith("RP-RM:"):
            allowed_formations = parse_rprm_formations(role)
            is_rprm = True
        elif role.startswith("RP-RM"):
            is_rprm = True

        data = build_parametrage_data(session, allowed_formations=allowed_formations)
        return templates.TemplateResponse(
            request=request,
            name="parametrage.html",
            context={
                "request": request,
                "templates": data["templates"],
                "campus_list": data["campusList"],
                "filieres": data["filieres"],
                "semestres": data["semestres"],
                "annees_scolaires": data["anneesScolaires"],
                "profs": data["profsList"],
                "ues_by_filiere": data["uesByFiliere"],
                "selected_template_id": data["selectedTemplateId"],
                "selected_campus_id": data["selectedCampusId"],
                "selected_filiere_id": data["selectedFiliereId"],
                "semestre_annee": data["semestreAnnee"],
                "selected_annee_scolaire": data["selectedAnneeScolaire"],
                "is_rprm": is_rprm,
            },
        )

    # └────────────────────────────────────────────────────────────────┘

    # ┌─ API : Données de paramétrage (accès restreint Admin + RP-RM) ────┐
    @app.get("/api/parametrage")
    def parametrage_api(request: Request, session: SessionDep):
        # ── Sécurité : vérifier que l'utilisateur est Admin ou RP-RM ──
        user = require_roles(request, ["Admin", "RP-RM"])
        if user is None:
            return JSONResponse(
                content={"error": "Accès refusé. Rôle Admin ou RP-RM requis."},
                status_code=403,
            )

        # Filtrer les filières pour les RP-RM
        allowed_formations = None
        role = user.get("role", "") or ""

        if ':' in role: #RP-RM or Admin with formations
            allowed_formations = parse_rprm_formations(role)

        return JSONResponse(
            content=build_parametrage_data(
                session, allowed_formations=allowed_formations
            )
        )

    # └────────────────────────────────────────────────────────────────┘

    # ┌─ API : Modules du sondage de l'année précédente ─────────────────┐
    @app.get("/api/modules-precedents")
    def modules_precedents_api(
        session: SessionDep,
        semestre: str = "",
        formation: str = "",
        annee_scolaire: str = "",
    ):
        """
        Retourne les modules du sondage de l'année scolaire précédente
        pour le même semestre et la même formation.

        Exemple : annee_scolaire="2025-2026" → cherche "2024-2025".
        Si aucun sondage précédent n'existe, renvoie un tableau vide.
        """
        if not semestre or not formation or not annee_scolaire:
            return JSONResponse(content={"ues": [], "profsList": []})

        # ── Calcul de l'année précédente ──────────────────────────────
        try:
            parts = annee_scolaire.split("-")
            if len(parts) == 2:
                year_start = int(parts[0]) - 1
                year_end = int(parts[1]) - 1
                annee_precedente = f"{year_start}-{year_end}"
            else:
                return JSONResponse(content={"ues": [], "profsList": []})
        except (ValueError, IndexError):
            return JSONResponse(content={"ues": [], "profsList": []})

        # ── Recherche du sondage de l'année précédente ────────────────
        sondage_precedent = session.exec(
            select(Sondage).where(
                Sondage.annee_scolaire == annee_precedente,
                Sondage.semestre == semestre,
                Sondage.formation == formation,
            )
        ).first()

        if not sondage_precedent:
            return JSONResponse(content={"ues": [], "profsList": []})

        # ── Récupération des modules liés à ce sondage ────────────────
        modules = session.exec(
            select(Module).where(
                Module.id_sondage == sondage_precedent.id_sondage,
                Module.id_template == sondage_precedent.id_template,
            )
        ).all()

        if not modules:
            return JSONResponse(content={"ues": [], "profsList": []})

        # ── Groupement par UE + extraction des professeurs ────────────
        ues_dict = {}
        seen_professors = {}
        professors = []
        professor_index = 1

        for module in modules:
            ue_name = module.ue or "Sans UE"
            if ue_name not in ues_dict:
                ues_dict[ue_name] = {
                    "id": len(ues_dict) + 1,
                    "nom": ue_name,
                    "optionnel": bool(module.ue_optionnelle),
                    "_open": True,
                    "modules": [],
                }

            prof_list = []
            if module.enseignant:
                prof_strings = [
                    p.strip() for p in module.enseignant.split(",") if p.strip()
                ]
                for prof_str in prof_strings:
                    parsed = parse_name(prof_str, professor_index)
                    if parsed["prenom"] or parsed["nom"]:
                        key = (
                            (parsed["prenom"] or "").lower(),
                            (parsed["nom"] or "").lower(),
                        )
                        if key not in seen_professors:
                            seen_professors[key] = professor_index
                            parsed["id"] = professor_index
                            professors.append(parsed)
                            professor_index += 1
                        prof_list.append(
                            {
                                "id": seen_professors[key],
                                "prenom": parsed["prenom"],
                                "nom": parsed["nom"],
                            }
                        )

            ues_dict[ue_name]["modules"].append(
                {
                    "id": int(module.id_module or 0),
                    "nom": module.nom or "Module",
                    "choix_enseignant_exclusif": bool(module.choix_enseignant),
                    "professeurs": prof_list,
                }
            )

        return JSONResponse(
            content={
                "ues": list(ues_dict.values()),
                "profsList": professors,
                "annee_precedente": annee_precedente,
                "sondage_id": sondage_precedent.id_sondage,
            }
        )

    # └────────────────────────────────────────────────────────────────┘

    # ┌─ API : Création d'un sondage (accès restreint Admin + RP-RM) ────┐
    @app.post("/api/sondage")
    async def create_sondage(
        request: Request,
        session: SessionDep,
        sondage_data: str = Form(...),
        file: Optional[UploadFile] = File(None),
    ):
        """
        Crée un sondage ET importe les étudiants en une seule transaction.
        Si l'import Excel échoue, le sondage est annulé (ROLLBACK).
        """
        # ── Sécurité : vérifier que l'utilisateur est Admin ou RP-RM ──
        user = require_roles(request, ["Admin", "RP-RM"])
        if user is None:
            return JSONResponse(
                content={"error": "Accès refusé. Rôle Admin ou RP-RM requis."},
                status_code=403,
            )

        # ── Parse le JSON du sondage envoyé en FormData ──
        try:
            sondage_dict = json.loads(sondage_data)
            sondage = SondageFullCreate(**sondage_dict)
        except Exception as e:
            return JSONResponse(
                content={"error": f"Données du sondage invalides : {str(e)}"},
                status_code=400,
            )

        # ── Sécurité : vérifier que la formation est autorisée pour le RP-RM ──
        role = user.get("role", "")
        if ':' in role: # RM-RP or Admin with formations
            allowed = parse_rprm_formations(role)
            if sondage.formation not in allowed:
                return JSONResponse(
                    content={
                        "error": f"Formation '{sondage.formation}' non autorisée pour votre rôle."
                    },
                    status_code=403,
                )

        # ── Pré-lecture du fichier Excel (avant la transaction) ──
        emails = []
        has_file = (
            file is not None
            and file.filename
            and file.filename.lower().endswith(".xlsx")
        )

        if file is not None and file.filename:
            if not file.filename.lower().endswith(".xlsx"):
                return JSONResponse(
                    content={
                        "error": "Format invalide. Seuls les fichiers .xlsx sont acceptés."
                    },
                    status_code=400,
                )

        if has_file:
            try:
                from openpyxl import load_workbook

                contents = await file.read()
                print(
                    f"[SONDAGE+IMPORT] Fichier reçu : {file.filename} ({len(contents)} octets)"
                )

                if len(contents) == 0:
                    return JSONResponse(
                        content={"error": "Le fichier Excel est vide."},
                        status_code=400,
                    )

                wb = load_workbook(filename=io.BytesIO(contents), read_only=True)
                ws = wb.active
                print(f"[SONDAGE+IMPORT] Feuille active : {ws.title}")

                row_count = 0
                for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
                    row_count += 1
                    cell_value = row[0]
                    if cell_value is None:
                        continue
                    cell_str = str(cell_value).strip()
                    if cell_str and "@" in cell_str:
                        emails.append(cell_str.lower())
                    elif cell_str and row_count <= 3:
                        print(
                            f"[SONDAGE+IMPORT] Ligne {row_count} ignorée (pas d'email) : '{cell_str}'"
                        )

                wb.close()
                print(
                    f"[SONDAGE+IMPORT] {row_count} ligne(s) lues, {len(emails)} email(s) valide(s)"
                )

                if not emails:
                    return JSONResponse(
                        content={
                            "error": f"Aucun email valide trouvé dans le fichier ({row_count} ligne(s) lue(s)). Vérifiez que les emails sont dans la première colonne."
                        },
                        status_code=400,
                    )

                # Dédupliquer
                emails = list(dict.fromkeys(emails))
                print(f"[SONDAGE+IMPORT] {len(emails)} email(s) unique(s) à traiter")

            except Exception as e:
                print(
                    f"[SONDAGE+IMPORT] Erreur lecture Excel : {type(e).__name__}: {e}"
                )
                return JSONResponse(
                    content={"error": f"Erreur de lecture du fichier Excel : {str(e)}"},
                    status_code=400,
                )

        # ── Transaction unique : Sondage + Modules + Users + Repondre ──
        nb_crees = 0
        nb_existants = 0
        nb_repondre_inseres = 0

        try:
            with session.begin():
                with session.no_autoflush:
                    # ── Étape 1 : Créer le sondage ──
                    existing_sondages = session.exec(
                        select(Sondage).where(
                            Sondage.id_template == sondage.id_template
                        )
                    ).all()
                    next_id_sondage = (
                        max([s.id_sondage for s in existing_sondages] + [0]) + 1
                    )

                    questionnaire_url = (
                        f"/questionnaire/{sondage.id_template}/{next_id_sondage}"
                    )

                    new_sondage = Sondage(
                        id_template=sondage.id_template,
                        id_sondage=next_id_sondage,
                        campus=sondage.campus,
                        formation=sondage.formation,
                        semestre=sondage.semestre,
                        annee_scolaire=sondage.annee_scolaire,
                        url=questionnaire_url,
                        statut=1,
                    )
                    session.add(new_sondage)

                    # ── Étape 2 : Créer les modules ──
                    for ue in sondage.ues:
                        for module_data in ue.modules:
                            prof_names = [
                                f"{p.prenom} {p.nom}" for p in module_data.professeurs
                            ]
                            enseignant_str = (
                                ", ".join(prof_names) if prof_names else None
                            )

                            new_module = Module(
                                nom=module_data.nom,
                                enseignant=enseignant_str,
                                ue=ue.nom,
                                ue_optionnelle=ue.optionnel,
                                choix_enseignant=module_data.choix_enseignant_exclusif,
                                id_template=sondage.id_template,
                                id_sondage=next_id_sondage,
                            )
                            session.add(new_module)

                    # ── Étape 3 : Importer les étudiants (si fichier fourni) ──
                    if emails:
                        email_to_user_id: Dict[str, int] = {}

                        existing_users = session.exec(select(User)).all()
                        existing_email_map = {
                            u.mail.lower(): u.id_user for u in existing_users if u.mail
                        }
                        print(
                            f"[SONDAGE+IMPORT] {len(existing_email_map)} user(s) existant(s) en BDD"
                        )

                        max_id = max([u.id_user for u in existing_users] + [0])

                        for email in emails:
                            if email in existing_email_map:
                                email_to_user_id[email] = existing_email_map[email]
                                nb_existants += 1
                            else:
                                max_id += 1
                                new_user = User(
                                    id_user=max_id,
                                    mail=email,
                                    role="Etudiant",
                                )
                                session.add(new_user)
                                email_to_user_id[email] = max_id
                                existing_email_map[email] = max_id
                                nb_crees += 1

                        print(
                            f"[SONDAGE+IMPORT] Users : {nb_crees} créé(s), {nb_existants} existant(s)"
                        )

                        user_ids = list(email_to_user_id.values())
                        if user_ids:
                            # Nettoyage préalable (DELETE) : Supprime toutes les lignes de la table Repondre
                            # où la valeur Id_User correspond à un des élèves présents
                            stmt = delete(Repondre).where(
                                Repondre.id_user.in_(user_ids)
                            )
                            res = session.exec(stmt)
                            print(
                                f"[SONDAGE+IMPORT] Nettoyage préalable : {res.rowcount} anciennes lignes supprimées de Repondre."
                            )

                            # Insertion (INSERT) : Uniquement après le nettoyage
                            for user_id in user_ids:
                                new_repondre = Repondre(
                                    id_template=sondage.id_template,
                                    id_sondage=next_id_sondage,
                                    id_user=user_id,
                                    repondu=False,
                                    date_soumission=None,
                                )
                                session.add(new_repondre)
                                nb_repondre_inseres += 1

                        print(
                            f"[SONDAGE+IMPORT] Repondre : {nb_repondre_inseres} inséré(s) (et nettoyés des conflits)"
                        )

            # Si on arrive ici, le COMMIT a été fait par le context manager
            print(f"[SONDAGE+IMPORT] Transaction COMMIT réussie !")

        except Exception as e:
            print(f"[SONDAGE+IMPORT] ERREUR — ROLLBACK : {type(e).__name__}: {e}")
            import traceback

            traceback.print_exc()
            return JSONResponse(
                content={"error": f"Erreur lors de la création du sondage : {str(e)}"},
                status_code=500,
            )

        result = {
            "message": "Sondage créé avec succès",
            "id_sondage": next_id_sondage,
            "questionnaire_url": questionnaire_url,
        }
        if emails:
            result.update(
                {
                    "nb_emails_lus": len(emails),
                    "nb_users_crees": nb_crees,
                    "nb_users_existants": nb_existants,
                    "nb_repondre_inseres": nb_repondre_inseres,
                }
            )
        return result

    # └────────────────────────────────────────────────────────────────┘

    # ┌─ Page questionnaire ─────────────────────────────────────────────┐
    @app.get("/questionnaire/{id_template}/{id_sondage}", response_class=HTMLResponse)
    def questionnaire_page(
        request: Request, id_template: int, id_sondage: int, session: SessionDep
    ):
        sondage = session.exec(
            select(Sondage).where(
                Sondage.id_template == id_template,
                Sondage.id_sondage == id_sondage,
            )
        ).first()
        if not sondage:
            return HTMLResponse(content="Sondage introuvable.", status_code=404)

        sections = session.exec(
            select(Section)
            .where(Section.id_template == id_template)
            .order_by(Section.ordre)
        ).all()
        questions = session.exec(
            select(Question).where(Question.id_template == id_template)
        ).all()
        options = session.exec(
            select(Option).where(Option.id_template == id_template)
        ).all()
        modules = session.exec(
            select(Module).where(Module.id_sondage == id_sondage)
        ).all()

        sections_data = []
        for sec in sections:
            sec_questions = [q for q in questions if q.id_section == sec.id_section]
            sec_questions.sort(key=lambda q: q.id_question)
            questions_data = []
            for q in sec_questions:
                q_options = [
                    o
                    for o in options
                    if o.id_section == sec.id_section and o.id_question == q.id_question
                ]
                q_options.sort(key=lambda o: o.id_option)
                questions_data.append(
                    {
                        "id_question": q.id_question,
                        "intitule": q.intitule,
                        "type": q.question_type,
                        "categorie": q.categorie,
                        "options": [
                            {"id_option": o.id_option, "intitule": o.intitule}
                            for o in q_options
                        ],
                    }
                )
            sections_data.append(
                {
                    "id_section": sec.id_section,
                    "nom": sec.nom,
                    "questions": questions_data,
                }
            )

        modules_data = []
        for mod in modules:
            profs = []
            if mod.enseignant:
                profs = [p.strip() for p in mod.enseignant.split(",") if p.strip()]
            modules_data.append(
                {
                    "id_module": mod.id_module,
                    "nom": mod.nom,
                    "ue": mod.ue,
                    "ue_optionnelle": bool(mod.ue_optionnelle),
                    "enseignants": profs,
                    "choix_enseignant": bool(mod.choix_enseignant),
                }
            )
        # Grouper les modules par UE pour la logique conditionnelle
        ues_data = {}
        for mod_data in modules_data:
            ue_name = mod_data["ue"] or "Sans UE"
            if ue_name not in ues_data:
                ues_data[ue_name] = {
                    "nom": ue_name,
                    "optionnelle": mod_data["ue_optionnelle"],
                    "modules": [],
                }
            ues_data[ue_name]["modules"].append(mod_data)
        ues_list = list(ues_data.values())

        return templates.TemplateResponse(
            request=request,
            name="questionnaire.html",
            context={
                "request": request,
                "sondage": {
                    "id_template": sondage.id_template,
                    "id_sondage": sondage.id_sondage,
                    "campus": sondage.campus,
                    "formation": sondage.formation,
                    "semestre": sondage.semestre,
                    "annee_scolaire": sondage.annee_scolaire,
                },
                "sections": sections_data,
                "modules": modules_data,
                "ues": ues_list,
            },
        )

    # └────────────────────────────────────────────────────────────────┘

    # ┌─ API : Soumission des réponses du questionnaire ─────────────────┐
    @app.post("/api/questionnaire/{id_template}/{id_sondage}/reponses")
    def submit_reponses(
        request: Request,
        id_template: int,
        id_sondage: int,
        submission: QuestionnaireSubmission,
        session: SessionDep,
    ):
        # 1. Authentification : récupérer l'utilisateur connecté (Azure Entra ID)
        user = get_current_user(request)
        if not user:
            return JSONResponse(
                content={"error": "Authentification requise. Veuillez vous connecter."},
                status_code=401,
            )

        # 2. Résoudre l'Id_User depuis l'email de l'utilisateur connecté
        db_user = session.exec(
            select(User).where(User.mail == user["email"].casefold())
        ).first()
        if not db_user:
            return JSONResponse(
                content={"error": "Utilisateur "+user["email"]+"non trouvé dans la base de données."},
                status_code=403,
            )

        # 3. Vérifier que le sondage existe
        sondage = session.exec(
            select(Sondage).where(
                Sondage.id_template == id_template,
                Sondage.id_sondage == id_sondage,
            )
        ).first()
        if not sondage:
            return JSONResponse(
                content={"error": "Sondage introuvable."}, status_code=404
            )

        # 4. Vérifier que cet élève est assigné à ce sondage (table Repondre)
        #    Règle stricte : pas de INSERT, UPDATE uniquement
        repondre = session.exec(
            select(Repondre).where(
                Repondre.id_template == id_template,
                Repondre.id_sondage == id_sondage,
                Repondre.id_user == db_user.id_user,
            )
        ).first()
        if not repondre:
            return JSONResponse(
                content={
                    "error": "Vous n'êtes pas autorisé ou assigné à répondre à ce sondage."
                },
                status_code=403,
            )

        # 5. Vérifier que l'élève n'a pas déjà soumis ses réponses
        if repondre.repondu:
            return JSONResponse(
                content={
                    "error": "Vous avez déjà soumis vos réponses pour ce sondage."
                },
                status_code=409,
            )

        # 6. Enregistrement atomique : insertion des réponses + UPDATE Repondre
        #    On utilise begin_nested() (SAVEPOINT) car la session a déjà une
        #    transaction implicite ouverte par le générateur get_session().
        try:
            with session.begin_nested():
                # Calculer le prochain Id_Reponse
                existing_reponses = session.exec(select(Reponse)).all()
                next_id_reponse = (
                    max([r.id_reponse for r in existing_reponses] + [0]) + 1
                )

                # Insérer chaque réponse individuelle dans la table Reponses
                for rep in submission.reponses:
                    new_reponse = Reponse(
                        id_template=id_template,
                        id_sondage=id_sondage,
                        id_template_1=id_template,
                        id_section=rep.id_section,
                        id_module=rep.module_id,
                        enseignant=rep.enseignant,
                        id_question=rep.id_question,
                        id_reponse=next_id_reponse,
                        valeur=rep.valeur,
                    )
                    session.add(new_reponse)
                    next_id_reponse += 1

                # UPDATE de la ligne Repondre : marquer comme répondu
                repondre.repondu = True
                repondre.date_soumission = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                session.add(repondre)

            # Commit de la transaction principale
            session.commit()

        except Exception as e:
            session.rollback()
            return JSONResponse(
                content={"error": f"Erreur lors de l'enregistrement : {str(e)}"},
                status_code=500,
            )

        return {
            "message": "Réponses enregistrées avec succès",
            "id_user": db_user.id_user,
        }

    # └────────────────────────────────────────────────────────────────┘

    # ┌─ Route : Dashboards par rôle ────────────────────────────────────┐

    @app.get("/dashboard/{role}", response_class=HTMLResponse)
    async def dashboard(request: Request, role: str, session: SessionDep):
        user = get_current_user(request)
        if not user:
            return RedirectResponse(url="/")
        if role not in VALID_ROLES:
            return RedirectResponse(url="/")

        user_slug = role_to_dashboard_slug(user.get("role", ""))
        if user_slug != "admin" and user_slug != role: # Admin can see all the pages
            return RedirectResponse(url=f"/dashboard/{user_slug}")

        template_map = {
            "admin": "dashboard/admin.html",
            "etudiant": "dashboard/etudiant.html",
            "rprm": "dashboard/RPRM.html",
        }

        filieres = []
        full_role = user.get("role", "")
        if ":" in full_role:
            filieres = [f.strip() for f in full_role.split(":", 1)[1].split(";") if f.strip()]

        context = {"user": user, "filieres": filieres}

        if role == "admin":
            db_users = session.exec(select(User)).all()
            context["users"] = [
                {"id_user": u.id_user, "mail": u.mail, "role": u.role} for u in db_users
            ]

        if role in ("admin", "rprm"):
            all_sondages = session.exec(select(Sondage)).all()
            if role == "rprm" and filieres:
                sondages_filtres = [s for s in all_sondages if s.formation in filieres]
            else:
                sondages_filtres = list(all_sondages)

            sondages_list = []
            for s in sondages_filtres:
                nb_inscrits = (
                    session.exec(
                        select(func.count(Repondre.id_user)).where(
                            Repondre.id_template == s.id_template,
                            Repondre.id_sondage == s.id_sondage,
                        )
                    ).first()
                    or 0
                )

                nb_repondants = (
                    session.exec(
                        select(func.count(Repondre.id_user)).where(
                            Repondre.id_template == s.id_template,
                            Repondre.id_sondage == s.id_sondage,
                            Repondre.repondu == True,
                        )
                    ).first()
                    or 0
                )

                sondages_list.append(
                    {
                        "id_template": s.id_template,
                        "id_sondage": s.id_sondage,
                        "campus": s.campus,
                        "formation": s.formation,
                        "semestre": s.semestre,
                        "annee_scolaire": s.annee_scolaire,
                        "url": s.url,
                        "nb_inscrits": nb_inscrits,
                        "nb_repondants": nb_repondants,
                    }
                )
            context["sondages"] = sondages_list

        return templates.TemplateResponse(
            request=request,
            name=template_map[role],
            context=context,
        )

    # └────────────────────────────────────────────────────────────────┘

    # ┌─ API : Questionnaire assigné à l'étudiant connecté ──────────────┐
    @app.get("/api/etudiant/questionnaire")
    def get_etudiant_questionnaire(request: Request, session: SessionDep):
        """
        Retourne le questionnaire assigné à l'étudiant connecté.
        Interroge la table Repondre pour récupérer id_sondage, id_template
        et le statut Repondu (0=False, 1=True).
        """
        user = get_current_user(request)
        if not user:
            return JSONResponse(
                content={"error": "Authentification requise."},
                status_code=401,
            )

        # Résoudre l'id_user depuis l'email
        db_user = session.exec(
            select(User).where(User.mail == user["email"].casefold())
        ).first()
        if not db_user:
            return JSONResponse(
                content={"error": "Utilisateur "+user["email"]+" non trouvé en base de données."},
                status_code=404,
            )

        # Chercher les entrées Repondre pour cet utilisateur
        repondre_entries = session.exec(
            select(Repondre).where(Repondre.id_user == db_user.id_user)
        ).all()

        if not repondre_entries:
            return JSONResponse(
                content={
                    "questionnaire": None,
                    "message": "Aucun questionnaire assigné.",
                },
                status_code=200,
            )

        # Prendre le premier questionnaire non répondu, sinon le dernier
        non_repondu = [r for r in repondre_entries if not r.repondu]
        entry = non_repondu[0] if non_repondu else repondre_entries[-1]

        # Récupérer les infos du sondage pour le contexte
        sondage = session.exec(
            select(Sondage).where(
                Sondage.id_template == entry.id_template,
                Sondage.id_sondage == entry.id_sondage,
            )
        ).first()

        return {
            "questionnaire": {
                "id_template": entry.id_template,
                "id_sondage": entry.id_sondage,
                "repondu": bool(entry.repondu),
                "url": f"/questionnaire/{entry.id_template}/{entry.id_sondage}",
                "formation": sondage.formation if sondage else None,
                "semestre": sondage.semestre if sondage else None,
            }
        }

    # └────────────────────────────────────────────────────────────────┘

    # ┌─ API : Gestion des rôles utilisateurs (accès restreint Admin) ────┐
    def _is_valid_role(role: str) -> bool:
        """Accepte 'Admin' ou 'Admin:filière1,filière2', 'Etudiant', 'RP-RM' ou 'RP-RM:filière1;filière2;...'"""
        if role in {"Etudiant"}:
            return True
        if role.startswith("Admin"):
            return True
        if role.startswith("RP-RM"):
            return True
        return False

    @app.get("/api/users")
    def get_users(request: Request, session: SessionDep):
        # ── Sécurité : seul un Admin peut lister tous les utilisateurs ──
        user = require_roles(request, ["Admin"])
        if user is None:
            return JSONResponse(
                content={"error": "Accès refusé. Rôle Admin requis."},
                status_code=403,
            )
        users = session.exec(select(User)).all()
        return [{"id_user": u.id_user, "mail": u.mail, "role": u.role} for u in users]

    @app.put("/api/users/{id_user}/role")
    def update_user_role(
        request: Request, id_user: int, body: RoleUpdate, session: SessionDep
    ):
        # ── Sécurité : seul un Admin peut modifier les rôles ──
        admin = require_roles(request, ["Admin"])
        if admin is None:
            return JSONResponse(
                content={"error": "Accès refusé. Rôle Admin requis."},
                status_code=403,
            )
        if not _is_valid_role(body.role):
            return JSONResponse(
                content={"detail": f"Rôle invalide : '{body.role}'"},
                status_code=422,
            )
        user = session.get(User, id_user)
        if not user:
            return JSONResponse(
                content={"detail": f"Utilisateur {id_user} introuvable"},
                status_code=404,
            )
        user.role = body.role
        session.add(user)
        session.commit()
        session.refresh(user)

        return {"id_user": user.id_user, "mail": user.mail, "role": user.role}

    # ┌─ Visualisation & Export CSV ──────────────────────────────────────┐
    def _check_sondage_access_and_status(
        session: Session,
        id_template: int,
        id_sondage: int,
        role: str,
        formations_autorisees: list[str],
    ):
        """Helper pour vérifier les accès et le statut de participation"""
        sondage = session.exec(
            select(Sondage).where(
                Sondage.id_template == id_template, Sondage.id_sondage == id_sondage
            )
        ).first()
        if not sondage:
            return None, {"error": "Sondage introuvable.", "status_code": 404}, None, None

        if role != "admin" and sondage.formation not in formations_autorisees:
            return None, {
                "error": f"Formation '{sondage.formation}' non autorisée pour votre rôle.",
                "status_code": 403,
            }, None, None

        nb_inscrits = (
            session.exec(
                select(func.count(Repondre.id_user)).where(
                    Repondre.id_template == id_template,
                    Repondre.id_sondage == id_sondage,
                )
            ).first()
            or 0
        )
        nb_repondants = (
            session.exec(
                select(func.count(Repondre.id_user)).where(
                    Repondre.id_template == id_template,
                    Repondre.id_sondage == id_sondage,
                    Repondre.repondu == True,
                )
            ).first()
            or 0
        )

        warning_msg = None
        if nb_repondants < nb_inscrits or sondage.statut == 1:
            warning_msg = f"Attention : Le sondage est toujours en cours. Seulement {nb_repondants} élève(s) ont répondu sur {nb_inscrits} inscrits."

        return sondage, warning_msg, nb_inscrits, nb_repondants

    @app.get("/api/export/{id_template}/{id_sondage}")
    def export_sondage_csv(
        request: Request, id_template: int, id_sondage: int, session: SessionDep
    ):
        user = require_roles(request, ["Admin", "RP-RM"])
        if user is None:
            return JSONResponse(content={"error": "Accès refusé."}, status_code=403)

        role = user.get("role", "") or ""
        formations_autorisees = (
            parse_rprm_formations(role) if ":" in role else []
        )
        admin_role = "admin" if role == "Admin" else "rprm"

        sondage, error_or_warning, _, _ = _check_sondage_access_and_status(
            session, id_template, id_sondage, admin_role, formations_autorisees
        )
        if not sondage:
            return JSONResponse(
                content={"error": error_or_warning["error"]},
                status_code=error_or_warning["status_code"],
            )

        # Utilisation de la BDD locale pour le loader sqlite3 natif
        sondage_obj = load_sondage_complet(
            "database/db_oceens.db", id_template, id_sondage
        )

        resp = generate_csv_response(sondage_obj)
        if isinstance(error_or_warning, str):
            resp.headers["X-Warning"] = "Sondage en cours - donnees partielles"

        return resp

    @app.get("/visualisation/{id_template}/{id_sondage}", response_class=HTMLResponse)
    def visualisation_page(
        request: Request, id_template: int, id_sondage: int, session: SessionDep
    ):
        user = require_roles(request, ["Admin", "RP-RM"])
        if user is None:
            return RedirectResponse(url="/")

        role = user.get("role", "") or ""
        formations_autorisees = (
            parse_rprm_formations(role) if role.startswith("RP-RM") else []
        )
        admin_role = "admin" if role == "Admin" else "rprm"

        sondage, error_or_warning, nb_inscrits, nb_repondants = (
            _check_sondage_access_and_status(
                session, id_template, id_sondage, admin_role, formations_autorisees
            )
        )
        if not sondage:
            return HTMLResponse(
                content=f"<h1>Erreur</h1><p>{error_or_warning['error']}</p>",
                status_code=error_or_warning["status_code"],
            )

        # Utilisation de la BDD locale pour le loader sqlite3 natif
        sondage_obj = load_sondage_complet(
            "database/db_oceens.db", id_template, id_sondage
        )

        viz_context = get_visualisation_context(sondage_obj)

        context = {
            "user": user,
            "sondage": sondage,
            "nb_inscrits": nb_inscrits,
            "nb_repondants": nb_repondants,
            "warning_msg": error_or_warning
            if isinstance(error_or_warning, str)
            else None,
            "viz_data": viz_context,
        }

        return templates.TemplateResponse(
            request=request, name="visualisation.html", context=context
        )

    @app.get("/visualisation/admin", response_class=HTMLResponse)
    def visualisation_admin_globale(request: Request, session: SessionDep):
        user = require_roles(request, ["Admin"])
        if user is None:
            return RedirectResponse(url="/")

        all_sondages = session.exec(select(Sondage)).all()

        # On regroupe les sondages par campus pour la vue de Niveau 0
        sondages_list = []
        for s in all_sondages:
            nb_inscrits = (
                session.exec(
                    select(func.count(Repondre.id_user)).where(
                        Repondre.id_template == s.id_template,
                        Repondre.id_sondage == s.id_sondage,
                    )
                ).first()
                or 0
            )

            nb_repondants = (
                session.exec(
                    select(func.count(Repondre.id_user)).where(
                        Repondre.id_template == s.id_template,
                        Repondre.id_sondage == s.id_sondage,
                        Repondre.repondu == True,
                    )
                ).first()
                or 0
            )

            sondages_list.append(
                {
                    "id_template": s.id_template,
                    "id_sondage": s.id_sondage,
                    "campus": s.campus,
                    "formation": s.formation,
                    "semestre": s.semestre,
                    "annee_scolaire": s.annee_scolaire,
                    "url": s.url,
                    "nb_inscrits": nb_inscrits,
                    "nb_repondants": nb_repondants,
                }
            )

        context = {"user": user, "sondages": sondages_list}
        return templates.TemplateResponse(
            request=request, name="visualisation_admin.html", context=context
        )

    # └───────────────────────────────────────────────────────────────────┘

    return app


# ┌─ Instance applicative globale ───────────────────────────────────────┐
app = create_app()
# └──────────────────────────────────────────────────────────────────────┘


if __name__ == "__main__":
    uvicorn.run(
        "app:app",
        host="0.0.0.0",
        port=8000,
        reload=False,
    )
